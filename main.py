import cv2
import time
import os
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook

# === 載入儲存好的人臉向量資料庫 ===
def load_face_database(data_folder='facedata'):
    database = {}
    for filename in os.listdir(data_folder):
        if filename.endswith('.npy'):
            name = os.path.splitext(filename)[0]
            vector = np.load(os.path.join(data_folder, filename))
            database[name] = vector
    return database

# === 臉部辨識（比對向量）===
from face_recognition import extract_face, get_embedding

def recognize_face(image, database, threshold=0.7):
    face = extract_face(image)
    if face is None:
        return "No face", None
    embedding = get_embedding(face)
    min_dist = float('inf')
    identity = "Unknown"
    for name, db_emb in database.items():
        dist = np.linalg.norm(embedding - db_emb)
        if dist < min_dist:
            min_dist = dist
            identity = name
    if min_dist > threshold:
        return "Unknown", min_dist
    return identity, min_dist

# === 訪問紀錄 Excel 初始化 ===
history_folder = "history"
record_file = os.path.join(history_folder, "access_log.xlsx")
os.makedirs(history_folder, exist_ok=True)

if not os.path.exists(record_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["日期", "時間", "姓名"])
    wb.save(record_file)
else:
    wb = load_workbook(record_file)
    ws = wb.active

def log_access(name):
    now = datetime.now()
    date_str = now.strftime("%Y%m%d")
    time_str = now.strftime("%H%M")
    ws.append([date_str, time_str, name])
    wb.save(record_file)

# === 主程式 ===
print("🔍 載入人臉向量資料庫...")
database = load_face_database('facedata')

if not database:
    print("❌ 沒有找到任何 .npy 向量檔案，請先執行預處理程式！")
    exit()

print("✅ 資料庫載入完成，啟動攝影機辨識中...")
cap = cv2.VideoCapture(0)

recognized_name = None
start_time = None
final_frame = None  # 保留最後畫面用於 unknown 截圖

while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame = cv2.flip(frame, 1)  # ✅ 左右鏡像修正
    final_frame = frame.copy()

    name, dist = recognize_face(frame, database)

    if name not in ["No face"]:
        current_required = 10 if name == "Unknown" else 3

        if recognized_name == name:
            elapsed = time.time() - start_time
            if name == "Unknown":
                label = f"❌ no famil detected {elapsed:.1f} s"
                color = (0, 0, 255)
            else:
                label = f"✅ {name} my famil detected {elapsed:.1f} s"
                color = (0, 255, 0)

            if elapsed >= current_required:
                print(f"🎉 {name} 持續存在 {current_required} 秒，自動關閉攝影機")
                log_access(name)

                if name == "Unknown":
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    img_path = os.path.join(history_folder, f"unknown_{timestamp}.png")
                    cv2.imwrite(img_path, final_frame)
                    print(f"📸 Unknown 截圖已儲存：{img_path}")
                break
        else:
            recognized_name = name
            start_time = time.time()
            if name == "Unknown":
                label = "❌ no famil（開始計時）"
                color = (0, 0, 255)
            else:
                label = f"✅ {name} 允許進入（開始計時）"
                color = (0, 255, 0)
    else:
        recognized_name = None
        start_time = None
        label = "no face"
        color = (128, 128, 128)

    cv2.putText(frame, label, (30, 40), cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)
    cv2.imshow("Access Control System", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        print("👋 手動結束辨識")
        break

cap.release()
cv2.destroyAllWindows()